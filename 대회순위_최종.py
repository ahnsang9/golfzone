from selenium import webdriver
from time import sleep
from datetime import datetime
import openpyxl

if __name__ == '__main__':
    def click(xpath):
        while 1:
            try:
                driver.find_element_by_xpath(xpath).click()
                break
            except:pass

    def crawling(info):
        rank = driver.find_elements_by_xpath('/html/body/div[2]/div[3]/div/div/div[3]/div[2]/div[3]/table/tbody/tr/td[1]/span')
        nickname = driver.find_elements_by_xpath('/html/body/div[2]/div[3]/div/div/div[3]/div[2]/div[3]/table/tbody/tr/td[2]/div/a[2]')
        round = driver.find_elements_by_xpath('/html/body/div[2]/div[3]/div/div/div[3]/div[2]/div[3]/table/tbody/tr/td[3]')
        a = driver.find_elements_by_xpath('/html/body/div[2]/div[3]/div/div/div[3]/div[2]/div[3]/table/tbody/tr/td[5]/a')
        b = driver.find_elements_by_xpath('/html/body/div[2]/div[3]/div/div/div[3]/div[2]/div[3]/table/tbody/tr/td[6]/a')
        score_correction = driver.find_elements_by_xpath('/html/body/div[2]/div[3]/div/div/div[3]/div[2]/div[3]/table/tbody/tr/td[7]/a')
        final_score = driver.find_elements_by_xpath('/html/body/div[2]/div[3]/div/div/div[3]/div[2]/div[3]/table/tbody/tr/td[8]')
        for i in range(len(rank)):
            info[0].append(rank[i].text.replace('T', '')+'위')
            #info[1].append((re.sub(r'\(.*?\)', '', nickname[i].text)).replace("\n",""))
            info[1].append((nickname[i].text).replace("\n", " "))
            info[2].append(round[i].text)
            info[3].append(a[i].text)
            info[4].append(b[i].text)
            info[5].append(score_correction[i].text)
            info[6].append(final_score[i].text)
        return info

    def delete(info_num, index):
        try:
            for k in range(7):
                info_num[k].pop(index)
        except: pass

    def rank(a):
        if info[1].index(info2[1][a]) <= 1 and a == 0:
            delete(info2, a)
        elif info[1].index(info2[1][a]) <= 1 and 0 < a <= 2:
            delete(info2, a)
        elif info[1].index(info2[1][a]) <= 4 and 2 < a <= 4:
            delete(info2, a)
        else:
            for k in range(info[1].index(info2[1][a]), 25):
                info[0][k] -= 1
            delete(info, info[1].index(info2[1][a]))


    driver_path = 'chromedriver.exe'
    url = 'http://m.golfzon.com/ghome/#!/tournament/1/14042'
    info = [[], [], [], [], [], [], []]   # 0.순위 1.닉네임 2.라운드수 3.A코스 4.B코스 5.스코어보정치 6.최종성적
    info2 = [[], [], [], [], [], [], []]  # 多라운드
    options = webdriver.ChromeOptions()
    options.add_argument('headless')    # headless 옵션 설정
    options.add_argument("no-sandbox")  # headless 옵션 설정
    driver = webdriver.Chrome(driver_path, options=options)
    driver.get(url)

    click('/html/body/div[1]/div/div[3]/div[2]/section/article/div[2]/ul')
    driver.switch_to.window(driver.window_handles[-1])
    sleep(1)
    crawling(info)

    comp_num = driver.find_element_by_xpath('/html/body/div[2]/div[3]/div/div/div[1]/dl/dt/strong').text[5:]
    title = comp_num + ' (최종순위)'
    title2 = comp_num + ' 多라운드 (최종순위)'

    try:
        if driver.find_element_by_xpath('/html/body/div[2]/div[3]/div/div/div[3]/div[2]/div[3]/div[1]/div/div/a[3]'):  # 다음페이지
            driver.find_element_by_xpath('/html/body/div[2]/div[3]/div/div/div[3]/div[2]/div[3]/div[1]/div/div/a[3]').click()
            sleep(1)
            crawling(info)
            if driver.find_element_by_xpath('/html/body/div[2]/div[3]/div/div/div[3]/div[2]/div[3]/div[1]/div/div/a[4]'):
                driver.find_element_by_xpath('/html/body/div[2]/div[3]/div/div/div[3]/div[2]/div[3]/div[1]/div/div/a[4]').click()
                sleep(1)
                crawling(info)
    except:
        pass

    click('/html/body/div[2]/div[3]/div/div/div[3]/div[2]/div[1]/ul/li[2]/input')
    sleep(1)
    crawling(info2)

    click('/html/body/div[2]/div[3]/div/div/div[2]/ul/li[2]/a')
    sleep(1)
    a_course = driver.find_element_by_xpath('/html/body/div[2]/div[3]/div/div/div[3]/table/tbody/tr[6]/td/ul/li[1]').text[5:]
    b_course = driver.find_element_by_xpath('/html/body/div[2]/div[3]/div/div/div[3]/table/tbody/tr[6]/td/ul/li[2]').text[5:]
    driver.quit()

    for i in range(len(info[0])):
        info[0][i] = int(info[0][i][:-1])

    if '스카이이글 (knl**)' in info[1]:
        for i in range(info[1].index('스카이이글 (knl**)'), 25):
            info[0][i] -= 1
        delete(info, info[1].index('스카이이글 (knl**)'))
    if '스카이이글 (knl**)' in info2[1]:
        delete(info2, info2[1].index('스카이이글 (knl**)'))

    while 1:
        try:
            for i in range(5):
                rank(i)
        except:
            break

    for i in range(len(info[0])):
        info[0][i] = str(info[0][i]) + '위'

    wb = openpyxl.load_workbook('매장대회.xlsx')
    temp = wb.active

    ################################################################# 순위
    for i in range(len(info)):
        for j in range(min(len(info[i]), 20)):
            temp.cell(row=j+4, column=i+1).value = info[i][j]

    ################################################################# 多라운드 순위
    temp1 = info2[2][:5]
    rank = [1,1,1,1,1]
    for i in range(4):
        rank[i+1] = i+2
        if int(temp1[i]) == int(temp1[i+1]):
            rank[i+1] = rank[i]
    for i in range(min(len(info2[1]), 5)):
        temp.cell(row=i+29, column=1).value = str(rank[i])+'위'
        temp.cell(row=i+29, column=2).value = info2[1][i]
        temp.cell(row=i+29, column=7).value = info2[2][i]

    temp.cell(row=1, column=1).value = title
    temp.cell(row=26, column=1).value = title2
    temp.cell(row=3, column=4).value = a_course
    temp.cell(row=3, column=5).value = b_course
    wb.save('매장대회(최종순위).xlsx')

    print(info)